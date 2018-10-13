#/usr/bin/env python
#coding=utf8
 
import hashlib 
import random
import openpyxl
from openpyxl import Workbook
import json
import urllib.request

print('translate begin...')

#set parameter
appid = '20180816000194959'
secretKey = 'lWzwUiWmhRORknf68FCT'
fromLang = 'en' # 源语言
toLang = 'zh'   # 翻译后的语言
srcFilename = 'c:\_Work\chinesetranslation.xlsx'
srcColumn=2
srcRowBegin=1
srcRowEnd=28
srcSheet = 'Sheet2'
desFilename = 'c:\_Work\chinesetranslation_result.xlsx'
desColumn=1
desSheet = 'result2'


def translateBaidu(content): 
    apiurl = 'http://api.fanyi.baidu.com/api/trans/vip/translate'
    q = content 
    salt = random.randint(32768, 65536)
    sign = str(appid) + q + str(salt) + secretKey 
    sign = hashlib.md5(sign.encode("utf-8")).hexdigest()
    myurl = apiurl + '?appid=' + appid + '&q=' + urllib.parse.quote(q) + '&from=' + fromLang + '&to=' + toLang + '&salt=' + str(salt) + '&sign=' + sign
 
    try: 
        paramas = {
            'q':q,
            'from':'zh',
            'to':'en',
            'appid':'%s'%appid,
            'salt':'%s'%salt,
            'sign':'%s'%sign
            } 
        response = urllib.request.urlopen(myurl)
        jsonResponse = response.read()# 获得返回的结果，结果为json格式
        js = json.loads(jsonResponse)  # 将json格式的结果转换字典结构
        dst = str(js["trans_result"][0]["dst"])  # 取得翻译后的文本结果 
        return dst
    except Exception as e:
        print(e) 


wb = openpyxl.load_workbook(srcFilename)
ws = wb[srcSheet]
wb2 = Workbook()
ws2 = wb2.create_sheet(title=desSheet)

for i in range (srcRowBegin,srcRowEnd,1):
    result=ws.cell(row=i, column=srcColumn).value
    if not (result is None):
        ws2.cell(row=i-srcRowBegin+1,column=desColumn).value = translateBaidu(result)

wb2.save(desFilename)
print('ending...') 